import imaplib
import email
import datetime
import io
import os
import pickle
from pathlib import Path
from email.header import decode_header
from dotenv import load_dotenv

# --- Excel Import ---
import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font 
from openpyxl.formatting.rule import CellIsRule 

# --- Database Import ---
import psycopg2

# --- Mindee Imports ---
from mindee import Client, product
from mindee.input import BytesInput

# --- Google Imports ---
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# Load environment variables from project root .env first, then local fallback
BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")
load_dotenv()

# ================= CONFIGURATION =================
# 📧 Email Config
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
IMAP_SERVER = os.getenv("IMAP_SERVER")
PROCESSED_LABEL = os.getenv("PROCESSED_LABEL")

# ☁️ Google Drive Config
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID")
CREDENTIALS_FILE = os.getenv("CREDENTIALS_FILE")
TOKEN_FILE = os.getenv("TOKEN_FILE")
SCOPES = ["https://www.googleapis.com/auth/drive"]

# 📊 Excel Config
EXCEL_FILE = os.getenv("EXCEL_FILE")

# 🗄️ PostgreSQL Config
DATABASE_URL = os.getenv("DATABASE_URL")
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS") or os.getenv("DB_PASSWORD")

# 🧠 Mindee Config 
MINDEE_V2_API_KEY = os.getenv("MINDEE_V2_API_KEY")
model_id = "1cd90980-2c6c-4d30-8952-af92c6db8786"
# 🔍 EXTENDED KEYWORDS
INVOICE_TERMS = [
    "invoice", "bill", "payment", "receipt", "total", "due",
    "order", "statement", "quote", "estimate", "contract", 
    "subscription", "purchase", "transaction", "amount", "inv"
]

# 📂 ALLOWED FILES
ALLOWED_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png')
# =================================================


# ---------- 1. SETUP MINDEE ----------
if not MINDEE_V2_API_KEY:
    print("❌ ERROR: Mindee API Key is missing. Check your .env file!")
    exit()

mindee_client = Client(api_key=MINDEE_V2_API_KEY)


# ---------- 2. GOOGLE DRIVE AUTH ----------
def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                raise FileNotFoundError(f"❌ {CREDENTIALS_FILE} not found.")
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as token:
            pickle.dump(creds, token)
    return build('drive', 'v3', credentials=creds)

drive_service = get_drive_service()


# ---------- 3. DRIVE UPLOAD & DOWNLOAD ----------
def upload_to_drive(file_bytes, filename):
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype="application/octet-stream", resumable=True)
    metadata = {"name": filename, "parents": [GOOGLE_DRIVE_FOLDER_ID]}
    uploaded = drive_service.files().create(
        body=metadata, 
        media_body=media, 
        fields="id, webViewLink"
    ).execute()
    return uploaded.get("id"), uploaded.get("webViewLink")

def download_from_drive(file_id):
    request = drive_service.files().get_media(fileId=file_id)
    file_buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, request)
    done = False
    while done is False:
        _, done = downloader.next_chunk()
    return file_buffer.getvalue()


# ---------- 4. MINDEE EXTRACTION ----------
def extract_invoice_data(file_bytes, filename):
    print(f"      🧠 Sending to Mindee (Standard Invoice API)...")
    try:
        input_source = BytesInput(file_bytes, filename)
        response = mindee_client.parse(product.InvoiceV4, input_source)
        prediction = response.document.inference.prediction

        def get_field_info(field):
            val = field.value if field and field.value is not None else "N/A"
            score = int(field.confidence * 100) if field and field.confidence else 0
            return val, score

        inv_val, inv_score = get_field_info(prediction.invoice_number)
        date_val, date_score = get_field_info(prediction.date)
        vendor_val, vendor_score = get_field_info(prediction.supplier_name)
        total_val, total_score = get_field_info(prediction.total_amount)
        tax_val, tax_score = get_field_info(prediction.total_tax)

        avg_score = int((inv_score + date_score + vendor_score + total_score) / 4)

        data = {
            "invoice_no": inv_val, "invoice_conf": inv_score,
            "date": date_val,      "date_conf": date_score,
            "vendor": vendor_val,  "vendor_conf": vendor_score,
            "amount": total_val,   "amount_conf": total_score,
            "tax": tax_val,        "tax_conf": tax_score,
            "avg_score": avg_score
        }
        return data
        
    except Exception as e:
        print(f"      ❌ Mindee Error: {e}")
        return None


# ---------- 5a. EXCEL LOGGING ----------
def save_to_excel(data, drive_link, filename):
    print(f"      📊 Saving to Excel ({EXCEL_FILE})...")
    
    headers = [
        "Timestamp", "Filename", "Doc Quality %", 
        "Vendor", "Vendor %", 
        "Invoice No", "Inv %", 
        "Date", "Date %", 
        "Total Amount", "Total %", 
        "Tax", "Drive Link"
    ]
    
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    row = [
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        filename,
        data["avg_score"],       
        data["vendor"], data["vendor_conf"],
        data["invoice_no"], data["invoice_conf"],
        data["date"], data["date_conf"],
        data["amount"], data["amount_conf"],
        data["tax"], 
        drive_link
    ]
    ws.append(row)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    percent_columns = ['C', 'E', 'G', 'I', 'K'] 
    
    current_row = ws.max_row
    for col in percent_columns:
        cell_ref = f"{col}{current_row}"
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator='lessThan', formula=['70'], stopIfTrue=True, fill=red_fill, font=red_font)
        )

    wb.save(EXCEL_FILE)


# ---------- 5b. POSTGRESQL LOGGING ----------
def save_to_postgres(data, drive_link, sender_email):
    print(f"      🗄️ Saving to PostgreSQL...")
    
    try:
        if DATABASE_URL:
            conn = psycopg2.connect(DATABASE_URL)
        else:
            conn = psycopg2.connect(
                host=DB_HOST,
                database=DB_NAME,
                user=DB_USER,
                password=DB_PASS
            )
        cur = conn.cursor()

        insert_query = """
            INSERT INTO invoices (
                invoice_no, vendor_name, date_of_invoice_created, 
                amount, tax, email_id, status, file_link
            ) VALUES (%s, %s, %s, %s, %s, %s, 'yet to start', %s)
            ON CONFLICT (invoice_no, vendor_name) DO NOTHING;
        """
        
        # Format date for SQL (handle N/A)
        inv_date = data["date"] if data["date"] != "N/A" else None
        
        # Format numbers for SQL (handle N/A and cast to float)
        try:
            amount = float(data["amount"]) if data["amount"] != "N/A" else 0.0
        except ValueError:
            amount = 0.0
            
        try:
            tax = float(data["tax"]) if data["tax"] != "N/A" else 0.0
        except ValueError:
            tax = 0.0

        cur.execute(insert_query, (
            str(data["invoice_no"])[:100],  # Prevents DB crash if string > 100 chars
            str(data["vendor"])[:255],      # Prevents DB crash if string > 255 chars
            inv_date,
            amount,
            tax,
            str(sender_email)[:255],        # Prevents DB crash if string > 255 chars
            drive_link
        ))

        conn.commit()
        cur.close()
        conn.close()
        print(f"      ✅ Successfully saved to database!")

    except Exception as e:
        print(f"      ❌ Database Error: {e}")


# ---------- 6. HELPER FUNCTIONS ----------
def decode_str(header_value):
    if not header_value: return ""
    decoded_list = decode_header(header_value)
    text_parts = []
    for content, encoding in decoded_list:
        if isinstance(content, bytes):
            text_parts.append(content.decode(encoding or "utf-8", errors="ignore"))
        else:
            text_parts.append(str(content))
    return "".join(text_parts)


# ---------- 7. MAIN LOGIC ----------
def connect_and_fetch():
    if not all([EMAIL_USER, EMAIL_PASS, IMAP_SERVER]):
        print("❌ ERROR: Email credentials are not loaded properly from .env")
        return

    print("Connecting to Gmail...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_USER, EMAIL_PASS)
    
    try:
        mail.create(PROCESSED_LABEL)
    except imaplib.IMAP4.error:
        pass 

    mail.select("INBOX")
    status, messages = mail.search(None, "UNSEEN")
    
    if status != "OK" or not messages[0]:
        print("No unread emails found.")
        mail.logout()
        return

    email_ids = messages[0].split()
    print(f"Found {len(email_ids)} unread emails")

    for e_id in email_ids:
        try:
            _, msg_data = mail.fetch(e_id, "(RFC822)")
            msg = email.message_from_bytes(msg_data[0][1])

            subject = decode_str(msg.get("Subject", ""))
            subject_lower = subject.lower()

            print(f"--------------------------------------------------")
            print(f"📧 Checking: '{subject}'")

            # 1. KEYWORD CHECK
            is_invoice_email = False
            for term in INVOICE_TERMS:
                if term in subject_lower:
                    is_invoice_email = True
                    print(f"   ✅ Keyword matched: '{term}'")
                    break
            
            if not is_invoice_email:
                print(f"   ❌ Skipped: Subject does not contain known invoice keywords.")
                continue 

            # 2. ATTACHMENT CHECK
            valid_attachments = []
            for part in msg.walk():
                content_disposition = str(part.get_content_disposition())
                
                if "attachment" in content_disposition or "inline" in content_disposition:
                    fname = part.get_filename()
                    if fname:
                        fname = decode_str(fname)
                        if fname.lower().endswith(ALLOWED_EXTENSIONS):
                            valid_attachments.append((fname, part))
                            print(f"   📎 Found File: {fname}")

            if not valid_attachments:
                print(f"   ❌ Skipped: No valid PDF/Image attachments found.")
                continue

            # 3. PROCESSING
            print(f"   🚀 Processing {len(valid_attachments)} file(s)...")
            
            for filename, part in valid_attachments:
                clean_name = "".join(c for c in filename if c.isalnum() or c in "._- ")
                new_name = f"{int(datetime.datetime.now().timestamp())}_{clean_name}"
                email_bytes = part.get_payload(decode=True)
                
                # UPLOAD TO DRIVE
                drive_id, drive_link = upload_to_drive(email_bytes, new_name)
                print(f"      ☁ Uploaded: {new_name}")

                # DOWNLOAD FROM DRIVE
                print(f"      📥 Downloading from Drive...")
                drive_bytes = download_from_drive(drive_id)

                # EXTRACT MINDEE DATA
                data = extract_invoice_data(drive_bytes, filename)
                
                if data:
                    print(f"      📊 Total: {data['amount']} (Conf: {data['amount_conf']}%)")
                    
                    # Log to Excel
                    save_to_excel(data, drive_link, new_name)

                    # Extract Sender Email
                    sender_email = decode_str(msg.get("From", ""))
                    
                    # Log to Database
                    save_to_postgres(data, drive_link, sender_email)

            # MOVE TO LABEL
            result = mail.copy(e_id, PROCESSED_LABEL)
            if result[0] == 'OK':
                mail.store(e_id, '+FLAGS', '\\Deleted')
                print(f"      🗄️  Moved to '{PROCESSED_LABEL}'")

        except Exception as e:
            print(f"   ❌ Error: {e}")

    mail.expunge()
    mail.logout()
    print("\nDone.")

if __name__ == "__main__":
    connect_and_fetch()