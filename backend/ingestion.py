import imaplib
import email
import datetime
import io
import os
import pickle
from email.header import decode_header
from dotenv import load_dotenv

# --- Excel ---
import openpyxl
from openpyxl import Workbook, load_workbook

# --- PostgreSQL ---
import psycopg2

# --- Mindee ---
from mindee import ClientV2, InferenceParameters, BytesInput

# --- Google Drive ---
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# =====================================================
# LOAD ENV
# =====================================================
load_dotenv()

EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
IMAP_SERVER = os.getenv("IMAP_SERVER")
PROCESSED_LABEL = os.getenv("PROCESSED_LABEL")

MINDEE_API_KEY = os.getenv("MINDEE_V2_API_KEY")
MODEL_ID = "1cd90980-2c6c-4d30-8952-af92c6db8786"

GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID")
CREDENTIALS_FILE = os.getenv("CREDENTIALS_FILE")
TOKEN_FILE = os.getenv("TOKEN_FILE")
SCOPES = ["https://www.googleapis.com/auth/drive"]

EXCEL_FILE = os.getenv("EXCEL_FILE")

DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS")

INVOICE_TERMS = [
    "invoice", "bill", "payment", "receipt", "total", "due",
    "order", "statement", "quote", "estimate", "contract",
    "subscription", "purchase", "transaction", "amount", "inv"
]

ALLOWED_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png')

# =====================================================
# GOOGLE DRIVE AUTH
# =====================================================
def get_drive_service():
    creds = None

    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                CREDENTIALS_FILE, SCOPES
            )
            creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, 'wb') as token:
            pickle.dump(creds, token)

    return build('drive', 'v3', credentials=creds)

drive_service = get_drive_service()

# =====================================================
# OCR EXTRACTION (MINDEE CUSTOM MODEL)
# =====================================================
def ocr_and_extract_data(file_name, file_bytes):

    mindee_client = ClientV2(MINDEE_API_KEY)

    params = InferenceParameters(
        model_id=MODEL_ID
    )

    input_source = BytesInput(
        file_bytes,
        filename=file_name
    )

    response = mindee_client.enqueue_and_get_inference(
        input_source, params
    )

    fields = response.inference.result.fields

    data = {
        "invoice_no": fields.get("invoice_number").value if fields.get("invoice_number") else None,
        "customer": fields.get("customer_name").value if fields.get("customer_name") else None,
        "date": fields.get("date").value if fields.get("date") else None,
        "vendor": fields.get("supplier_name").value if fields.get("supplier_name") else None,
        "po_no": fields.get("po_number").value if fields.get("po_number") else None,
        "amount": fields.get("total_amount").value if fields.get("total_amount") else None,
        "tax": fields.get("total_tax").value if fields.get("total_tax") else None,
    }

    print("      🧠 Extracted:", data)
    return data

# =====================================================
# DRIVE UPLOAD
# =====================================================
def upload_to_drive(file_bytes, filename):

    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype="application/octet-stream",
        resumable=True
    )

    metadata = {
        "name": filename,
        "parents": [GOOGLE_DRIVE_FOLDER_ID]
    }

    uploaded = drive_service.files().create(
        body=metadata,
        media_body=media,
        fields="id, webViewLink"
    ).execute()

    return uploaded.get("id"), uploaded.get("webViewLink")

# =====================================================
# EXCEL LOGGING
# =====================================================
def save_to_excel(filename, drive_link):

    headers = ["Timestamp", "Filename", "Drive Link", "Status"]

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        filename,
        drive_link,
        "Uploaded"
    ])

    wb.save(EXCEL_FILE)

# =====================================================
# POSTGRESQL SAVE
# =====================================================
def save_to_postgres(data, drive_link):

    try:
        conn = psycopg2.connect(
            host=DB_HOST,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASS
        )

        cur = conn.cursor()

        query = """
        INSERT INTO invoice (
            invoice_no,
            customer_name,
            vendor_name,
            po_number,
            invoice_date,
            total_amount,
            total_tax,
            file_link
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s);
        """

        # Safe conversions
        try:
            amount = float(data["amount"]) if data["amount"] else 0.0
        except:
            amount = 0.0

        try:
            tax = float(data["tax"]) if data["tax"] else 0.0
        except:
            tax = 0.0

        cur.execute(query, (
            str(data["invoice_no"])[:100] if data["invoice_no"] else None,
            str(data["customer"])[:255] if data["customer"] else None,
            str(data["vendor"])[:255] if data["vendor"] else None,
            str(data["po_no"])[:100] if data["po_no"] else None,
            data["date"],
            amount,
            tax,
            drive_link
        ))

        conn.commit()
        cur.close()
        conn.close()

        print("      🗄️ Saved to PostgreSQL")

    except Exception as e:
        print("      ❌ Database Error:", e)

# =====================================================
# EMAIL PROCESSING
# =====================================================
def decode_str(header_value):
    if not header_value:
        return ""
    decoded_list = decode_header(header_value)
    text_parts = []
    for content, encoding in decoded_list:
        if isinstance(content, bytes):
            text_parts.append(content.decode(encoding or "utf-8", errors="ignore"))
        else:
            text_parts.append(str(content))
    return "".join(text_parts)

def connect_and_fetch():

    print("Connecting to Gmail...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_USER, EMAIL_PASS)

    try:
        mail.create(PROCESSED_LABEL)
    except:
        pass

    mail.select("INBOX")
    status, messages = mail.search(None, "UNSEEN")

    if status != "OK" or not messages[0]:
        print("No unread emails found.")
        return

    for e_id in messages[0].split():

        _, msg_data = mail.fetch(e_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])

        subject = decode_str(msg.get("Subject", ""))
        subject_lower = subject.lower()

        if not any(term in subject_lower for term in INVOICE_TERMS):
            continue

        for part in msg.walk():

            fname = part.get_filename()
            if fname and fname.lower().endswith(ALLOWED_EXTENSIONS):

                file_bytes = part.get_payload(decode=True)

                clean_name = "".join(c for c in fname if c.isalnum() or c in "._- ")
                new_name = f"{int(datetime.datetime.now().timestamp())}_{clean_name}"

                # OCR
                data = ocr_and_extract_data(fname, file_bytes)

                # Drive Upload
                drive_id, drive_link = upload_to_drive(file_bytes, new_name)
                print(f"      ☁ Uploaded: {new_name}")

                # Excel
                save_to_excel(new_name, drive_link)

                # PostgreSQL
                if data:
                    save_to_postgres(data, drive_link)

        mail.copy(e_id, PROCESSED_LABEL)
        mail.store(e_id, '+FLAGS', '\\Deleted')

    mail.expunge()
    mail.logout()
    print("Done.")

# =====================================================
if __name__ == "__main__":
    connect_and_fetch()