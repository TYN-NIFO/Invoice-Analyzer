import imaplib
import email
import datetime
import io
import json
import os
import pickle
import time
from email.header import decode_header
from mindee import ClientV2, InferenceParameters, BytesInput
from sqlalchemy.orm import Session
from models import Invoice, LineItem, EmailIngestionLog
from database import SessionLocal
import logging
from typing import List, Tuple, Dict
from config import (
    EMAIL_USER, EMAIL_PASS, IMAP_SERVER, PROCESSED_LABEL,
    MINDEE_API_KEY, MINDEE_MODEL_ID,
    INVOICE_TERMS, ALLOWED_EXTENSIONS, EXCEL_FILE
)
import boto3
import uuid
from botocore.config import Config as BotoConfig

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

STATE_FILE = os.path.join(os.path.dirname(__file__), "email_ingestion_state.json")

def parse_invoice_date(value) -> datetime.datetime:
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)

    if value is None:
        return datetime.datetime.now()

    date_text = str(value).strip()
    if not date_text:
        return datetime.datetime.now()

    try:
        return datetime.datetime.fromisoformat(date_text)
    except Exception:
        pass

    supported_formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m-%d-%Y",
        "%m/%d/%Y",
    ]
    for date_format in supported_formats:
        try:
            return datetime.datetime.strptime(date_text, date_format)
        except Exception:
            continue

    return datetime.datetime.now()

def sanitize_filename(filename: str) -> str:
    clean_name = "".join(c for c in filename if c.isalnum() or c in "._- ")
    return f"{int(datetime.datetime.now().timestamp())}_{clean_name}"

def save_invoice_to_excel(ocr_data: Dict, drive_link: str, filename: str, source: str = "email_ingestion") -> Tuple[bool, str]:
    """Append invoice extraction summary to Excel log file."""
    try:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            logger.warning("[EXCEL] openpyxl is not installed; skipping Excel sheet update")
            return False, "openpyxl is not installed"

        headers = [
            "Timestamp", "Source", "Filename", "Vendor", "Invoice Number",
            "Invoice Date", "PO Number", "Amount", "Tax", "Total Amount", "Drive Link"
        ]

        if not os.path.exists(EXCEL_FILE):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(headers)
        else:
            workbook = load_workbook(EXCEL_FILE)
            worksheet = workbook.active

        worksheet.append([
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            source,
            filename,
            ocr_data.get("vendor_name", "N/A"),
            ocr_data.get("invoice_number", "N/A"),
            ocr_data.get("date", "N/A"),
            ocr_data.get("po_number", "N/A"),
            float(ocr_data.get("amount", 0) or 0),
            float(ocr_data.get("tax", 0) or 0),
            float(ocr_data.get("total_amount", 0) or 0),
            drive_link or ""
        ])

        workbook.save(EXCEL_FILE)
        return True, ""
    except Exception as excel_error:
        logger.warning(f"[EXCEL] Failed to update Excel log: {excel_error}")
        return False, str(excel_error)

def process_manual_invoice_upload(file_name: str, file_bytes: bytes, uploaded_by: str = "admin") -> Dict:
    """Process manually uploaded invoice using same flow as email ingestion."""
    if not file_name.lower().endswith(ALLOWED_EXTENSIONS):
        raise ValueError("Unsupported file type for invoice upload")

    ocr_data = ocr_and_extract_data(file_name, file_bytes)
    if not ocr_data:
        raise ValueError("OCR extraction failed")

    invoice_number = ocr_data.get("invoice_number")
    if invoice_number and str(invoice_number).strip() and str(invoice_number).strip() != "N/A":
        db = SessionLocal()
        try:
            existing_invoice = db.query(Invoice).filter(Invoice.invoice_number == str(invoice_number).strip()).first()
            if existing_invoice:
                raise ValueError(f"Duplicate invoice detected: {invoice_number} already exists")
        finally:
            db.close()

    drive_file_name = sanitize_filename(file_name)
    drive_id, drive_link = upload_to_drive(file_bytes, drive_file_name)
    if not drive_id or not drive_link:
        raise ValueError("Google Drive upload failed")

    invoice_id = save_invoice_to_db(
        ocr_data=ocr_data,
        drive_file_id=drive_id,
        drive_link=drive_link,
        email_subject=f"Manual Upload by {uploaded_by}",
        file_name=file_name,
        pdf_url=drive_link,
        allow_update_existing=True
    )

    if not invoice_id:
        raise ValueError("Failed to save invoice to database")

    now = datetime.datetime.now()
    log_ok = log_ingestion(
        email_subject=f"Manual Upload by {uploaded_by}",
        filename=file_name,
        email_from=uploaded_by,
        email_date=now,
        status="success",
        drive_file_id=drive_id,
        drive_link=drive_link,
        invoice_id=invoice_id
    )
    if not log_ok:
        raise ValueError("Failed to write ingestion log")

    excel_ok, excel_error = save_invoice_to_excel(
        ocr_data=ocr_data,
        drive_link=drive_link,
        filename=file_name,
        source="manual_upload"
    )
    if not excel_ok:
        raise ValueError(f"Failed to update Excel sheet: {excel_error}")

    return {
        "invoice_id": invoice_id,
        "invoice_number": ocr_data.get("invoice_number"),
        "vendor_name": ocr_data.get("vendor_name"),
        "total_amount": ocr_data.get("total_amount", 0),
        "drive_file_id": drive_id,
        "drive_link": drive_link
    }

# ================= S3 STORAGE =================

AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY", "")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_KEY", "")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
BUCKET_NAME = os.getenv("BUCKET_NAME", "tyn-claims-app-storage-prod")
S3_PREFIX = "invoice-analyzer/"

_s3_client = None
def get_s3_client():
    global _s3_client
    if _s3_client is None and AWS_ACCESS_KEY and AWS_SECRET_KEY:
        _s3_client = boto3.client(
            's3',
            aws_access_key_id=AWS_ACCESS_KEY,
            aws_secret_access_key=AWS_SECRET_KEY,
            region_name=AWS_REGION,
            config=BotoConfig(connect_timeout=10, read_timeout=30, retries={'max_attempts': 2})
        )
    return _s3_client

# ================= OCR EXTRACTION =================

def ocr_and_extract_data(file_name: str, file_bytes: bytes) -> Dict:
    """Extract invoice data using Mindee OCR with retry logic"""
    max_retries = 3
    try:
        for attempt in range(max_retries):
            try:
                mindee_client = ClientV2(api_key=MINDEE_API_KEY)
                params = InferenceParameters(
                    model_id=MINDEE_MODEL_ID,
                    rag=None,
                    raw_text=None,
                    polygon=None,
                    confidence=None,
                )
                
                input_source = BytesInput(file_bytes, filename=file_name)
                response = mindee_client.enqueue_and_get_inference(input_source, params)
                fields = response.inference.result.fields
                break
            except Exception as inner_e:
                if attempt < max_retries - 1:
                    logger.warning(f"OCR attempt {attempt+1} failed, retrying in 3s: {str(inner_e)}")
                    time.sleep(3)
                else:
                    raise
        
        # Extract date with proper None handling
        date_value = None
        if "date" in fields and fields["date"].value:
            date_value = str(fields["date"].value)
        else:
            date_value = datetime.datetime.now().isoformat()
        
        data = {
            "invoice_number": fields.get("invoice_number", {}).value if "invoice_number" in fields else "N/A",
            "customer_name": fields.get("customer_name", {}).value if "customer_name" in fields else "N/A",
            "date": date_value,
            "vendor_name": fields.get("supplier_name", {}).value if "supplier_name" in fields else "N/A",
            "po_number": fields.get("po_number", {}).value if "po_number" in fields else "N/A",
            "amount": float(fields.get("total_amount", {}).value or 0) if "total_amount" in fields else 0.0,
            "tax": float(fields.get("total_tax", {}).value or 0) if "total_tax" in fields else 0.0,
        }
        
        line_items = []
        if "line_items" in fields and fields["line_items"].items:
            for item in fields["line_items"].items:
                sub = item.fields
                line_items.append({
                    "description": sub.get("description", {}).value if "description" in sub else "N/A",
                    "quantity": float(sub.get("quantity", {}).value or 0) if "quantity" in sub else 0,
                    "unit_price": float(sub.get("unit_price", {}).value or 0) if "unit_price" in sub else 0,
                    "total_price": float(sub.get("total_price", {}).value or 0) if "total_price" in sub else 0,
                })
        
        data["line_items"] = line_items
        data["total_amount"] = data["amount"] + data["tax"]
        
        return data
    except Exception as e:
        logger.error(f"OCR error for {file_name}: {str(e)}")
        return None

# ================= S3 UPLOAD =================

def upload_to_drive(file_bytes: bytes, filename: str) -> Tuple[str, str]:
    """Upload file to S3 and return (file_id, url)"""
    try:
        s3 = get_s3_client()
        if not s3:
            logger.error("[S3] S3 not configured (missing AWS credentials)")
            return "no-s3", ""

        file_id = str(uuid.uuid4())
        s3_key = f"{S3_PREFIX}{file_id}/{filename}"

        s3.put_object(
            Bucket=BUCKET_NAME,
            Key=s3_key,
            Body=file_bytes,
            ContentType="application/octet-stream"
        )

        url = f"https://{BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{s3_key}"
        logger.info(f"[S3] Uploaded {filename} to {s3_key}")
        return file_id, url
    except Exception as e:
        logger.error(f"S3 upload error for {filename}: {str(e)}")
        return None, None

# ================= DATABASE OPERATIONS =================

def save_invoice_to_db(
    ocr_data: Dict,
    drive_file_id: str,
    drive_link: str,
    email_subject: str,
    file_name: str,
    pdf_url: str = None,
    allow_update_existing: bool = False
) -> int:
    """Save invoice and line items to database"""
    db = SessionLocal()
    try:
        invoice_number = ocr_data.get("invoice_number", f"INV-{datetime.datetime.now().timestamp()}")

        existing_invoice = db.query(Invoice).filter(Invoice.invoice_number == invoice_number).first()
        if existing_invoice:
            if allow_update_existing:
                existing_invoice.vendor_name = ocr_data.get("vendor_name", existing_invoice.vendor_name)
                existing_invoice.customer_name = ocr_data.get("customer_name", existing_invoice.customer_name)
                existing_invoice.po_number = ocr_data.get("po_number", existing_invoice.po_number)
                existing_invoice.invoice_date = parse_invoice_date(ocr_data.get("date"))
                existing_invoice.amount = ocr_data.get("amount", existing_invoice.amount)
                existing_invoice.tax = ocr_data.get("tax", existing_invoice.tax)
                existing_invoice.total_amount = ocr_data.get("total_amount", existing_invoice.total_amount)
                existing_invoice.email_subject = email_subject
                existing_invoice.pdf_url = pdf_url or drive_link
                existing_invoice.drive_file_id = drive_file_id
                existing_invoice.ocr_data = ocr_data
                existing_invoice.status = "pending"
                existing_invoice.reviewed_by = None
                existing_invoice.reviewed_at = None
                existing_invoice.updated_at = datetime.datetime.now()

                db.query(LineItem).filter(LineItem.invoice_id == existing_invoice.id).delete()
                for item in ocr_data.get("line_items", []):
                    db.add(LineItem(
                        invoice_id=existing_invoice.id,
                        description=item.get("description", ""),
                        quantity=item.get("quantity", 0),
                        unit_price=item.get("unit_price", 0),
                        total_price=item.get("total_price", 0)
                    ))

                db.commit()
                return existing_invoice.id
            return existing_invoice.id

        # Create invoice
        invoice = Invoice(
            invoice_number=invoice_number,
            vendor_name=ocr_data.get("vendor_name", "Unknown"),
            customer_name=ocr_data.get("customer_name", "Unknown"),
            po_number=ocr_data.get("po_number"),
            invoice_date=parse_invoice_date(ocr_data.get("date")),
            amount=ocr_data.get("amount", 0),
            tax=ocr_data.get("tax", 0),
            total_amount=ocr_data.get("total_amount", 0),
            status="pending",
            email_subject=email_subject,
            pdf_url=pdf_url or drive_link,
            drive_file_id=drive_file_id,
            ocr_data=ocr_data
        )
        
        db.add(invoice)
        db.flush()
        
        # Add line items
        for item in ocr_data.get("line_items", []):
            line_item = LineItem(
                invoice_id=invoice.id,
                description=item.get("description", ""),
                quantity=item.get("quantity", 0),
                unit_price=item.get("unit_price", 0),
                total_price=item.get("total_price", 0)
            )
            db.add(line_item)
        
        db.commit()
        return invoice.id
    except Exception as e:
        db.rollback()
        logger.error(f"Database error saving invoice: {str(e)}")
        return None
    finally:
        db.close()

def log_ingestion(
    email_subject: str,
    filename: str,
    email_from: str,
    email_date: datetime.datetime,
    status: str,
    drive_file_id: str = None,
    drive_link: str = None,
    error_message: str = None,
    invoice_id: int = None
):
    """Log email ingestion attempt"""
    db = SessionLocal()
    try:
        log = EmailIngestionLog(
            email_subject=email_subject,
            filename=filename,
            email_from=email_from,
            email_date=email_date,
            status=status,
            invoice_id=invoice_id,
            drive_file_id=drive_file_id,
            drive_link=drive_link,
            error_message=error_message
        )
        db.add(log)
        db.commit()
        return True
    except Exception as e:
        logger.error(f"Error logging ingestion: {str(e)}")
        return False
    finally:
        db.close()

# ================= HELPER FUNCTIONS =================

def decode_str(header_value: str) -> str:
    """Decode email header string"""
    if not header_value:
        return ""
    decoded_list = decode_header(header_value)
    text_parts = []
    for content, encoding in decoded_list:
        if isinstance(content, bytes):
            text_parts.append(content.decode(encoding or "utf-8", errors="ignore"))
        else:
            text_parts.append(str(content))
    return "".join(text_parts)


def is_already_logged(email_subject: str, filename: str, email_from: str, email_date: datetime.datetime) -> bool:
    """Check if an email attachment was already ingested/logged"""
    db = SessionLocal()
    try:
        existing = db.query(EmailIngestionLog).filter(
            EmailIngestionLog.email_subject == email_subject,
            EmailIngestionLog.filename == filename,
            EmailIngestionLog.email_from == email_from,
            EmailIngestionLog.email_date == email_date,
        ).first()
        return existing is not None
    finally:
        db.close()


def load_last_processed_uid() -> int:
    """Load last processed IMAP UID from local state file."""
    if not os.path.exists(STATE_FILE):
        return 0
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as state_file:
            data = json.load(state_file)
        return int(data.get("last_processed_uid", 0) or 0)
    except Exception as state_error:
        logger.warning(f"[STATE] Could not read state file: {state_error}")
        return 0


def save_last_processed_uid(uid_value: int) -> None:
    """Persist last processed IMAP UID to local state file."""
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as state_file:
            json.dump({"last_processed_uid": int(uid_value)}, state_file)
    except Exception as state_error:
        logger.warning(f"[STATE] Could not write state file: {state_error}")

# ================= MAIN EMAIL PROCESSING =================

def connect_and_fetch():
    """Connect to Gmail, fetch emails, process invoices"""
    try:
        logger.info("[*] Connecting to Gmail...")
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASS)
        
        # Create processed label if doesn't exist
        try:
            mail.create(PROCESSED_LABEL)
        except imaplib.IMAP4.error:
            pass
        
        mail.select("INBOX")
        since_date = (datetime.datetime.now() - datetime.timedelta(days=2)).strftime("%d-%b-%Y")
        last_processed_uid = load_last_processed_uid()

        if last_processed_uid > 0:
            status, messages = mail.uid("search", None, f"UID {last_processed_uid + 1}:*")
            logger.info(f"[EMAIL] Incremental UID scan from {last_processed_uid + 1}")
        else:
            status, messages = mail.uid("search", None, "SINCE", since_date)
            logger.info(f"[EMAIL] Initial UID scan using SINCE {since_date}")
        
        if status != "OK" or not messages[0]:
            logger.info("[OK] No recent emails found.")
            mail.logout()
            return {"message": "No emails to process", "count": 0}
        
        email_ids = messages[0].split()
        logger.info(f"[EMAIL] Found {len(email_ids)} candidate emails")
        
        processed_count = 0
        max_seen_uid = last_processed_uid
        
        for e_id in email_ids:
            try:
                # Decode email ID if it's bytes
                if isinstance(e_id, bytes):
                    e_id_str = e_id.decode()
                else:
                    e_id_str = str(e_id)
                
                _, msg_data = mail.uid("fetch", e_id_str, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                
                subject = decode_str(msg.get("Subject", ""))
                email_from = decode_str(msg.get("From", ""))
                email_date = email.utils.parsedate_to_datetime(msg.get("Date", str(datetime.datetime.now())))
                subject_lower = subject.lower()
                
                logger.info(f"[EMAIL] Processing: '{subject}'")
                
                # Check keywords
                is_invoice_email = any(term in subject_lower for term in INVOICE_TERMS)
                if not is_invoice_email:
                    logger.info("   [SKIP] No invoice keywords")
                    if not is_already_logged(subject, "N/A", email_from, email_date):
                        log_ingestion(subject, "N/A", email_from, email_date, "skipped", error_message="No invoice keywords")
                    continue
                
                # Check attachments
                valid_attachments = []
                for part in msg.walk():
                    content_disposition = str(part.get_content_disposition())
                    if "attachment" in content_disposition or "inline" in content_disposition:
                        fname = part.get_filename()
                        if fname and fname.lower().endswith(ALLOWED_EXTENSIONS):
                            valid_attachments.append((fname, part))
                
                if not valid_attachments:
                    logger.info("   [SKIP] No valid attachments")
                    if not is_already_logged(subject, "N/A", email_from, email_date):
                        log_ingestion(subject, "N/A", email_from, email_date, "skipped", error_message="No valid attachments")
                    continue
                
                # Process attachments
                for filename, part in valid_attachments:
                    try:
                        if is_already_logged(subject, filename, email_from, email_date):
                            logger.info(f"   [SKIP] Already processed: {filename}")
                            continue

                        email_bytes = part.get_payload(decode=True)
                        
                        # OCR Extract
                        ocr_data = ocr_and_extract_data(filename, email_bytes)
                        if not ocr_data:
                            raise Exception("OCR extraction failed")
                        
                        # Upload to Drive
                        clean_name = "".join(c for c in filename if c.isalnum() or c in "._- ")
                        new_name = f"{int(datetime.datetime.now().timestamp())}_{clean_name}"
                        drive_id, drive_link = upload_to_drive(email_bytes, new_name)
                        
                        # Save to DB
                        invoice_id = save_invoice_to_db(
                            ocr_data, drive_id, drive_link, subject, filename, pdf_url=drive_link
                        )
                        if not invoice_id:
                            raise Exception("Failed to save invoice to database")
                        
                        # Log success
                        log_ingestion(
                            subject, filename, email_from, email_date,
                            "success", drive_id, drive_link, invoice_id=invoice_id
                        )
                        
                        logger.info(f"   [OK] Processed: {filename}")
                        processed_count += 1
                        
                    except Exception as e:
                        logger.error(f"   [ERROR] {str(e)}")
                        log_ingestion(subject, filename, email_from, email_date, "failed", error_message=str(e))
                
                # Mark as read (DO NOT DELETE)
                try:
                    mail.store(str(e_id_str), '+FLAGS', '\\Seen')
                except Exception as flag_error:
                    logger.warning(f"   [WARN] Could not mark as read: {str(flag_error)}")

                # Optionally copy to processed label
                try:
                    mail.copy(str(e_id_str), PROCESSED_LABEL)
                except Exception:
                    pass  # Label might not exist

                try:
                    max_seen_uid = max(max_seen_uid, int(e_id_str))
                except Exception:
                    pass
                
            except Exception as e:
                logger.error(f"Email processing error: {str(e)}")

                try:
                    max_seen_uid = max(max_seen_uid, int(e_id_str))
                except Exception:
                    pass

        if max_seen_uid > last_processed_uid:
            save_last_processed_uid(max_seen_uid)
            logger.info(f"[STATE] Updated last_processed_uid to {max_seen_uid}")
        
        # DO NOT expunge - keep all emails
        mail.logout()
        
        return {
            "message": f"Processing complete",
            "processed": processed_count,
            "total": len(email_ids)
        }
        
    except Exception as e:
        logger.error(f"Fatal error: {str(e)}")
        return {"error": str(e)}

# Async wrapper for API endpoint
async def process_emails_async():
    """Async wrapper for email processing"""
    return connect_and_fetch()
